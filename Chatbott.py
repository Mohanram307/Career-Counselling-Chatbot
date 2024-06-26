import json
import openpyxl  # Library for working with Excel files

# Questions from the RIASEC Career Interest Quiz
questions = {
    "Interst Questions": [
        "I would rather:\n(R) Build a piece of furniture from scratch.\n(I) Conduct a scientific experiment.\n(A) Write a song or poem.\n(S) Help others solve their problems.\n(E) Lead a team project.\n(C) Follow clear instructions to complete a task.",
        "In my free time, I enjoy:\n(R) Working with tools and fixing things.\n(I) Researching new topics and learning new things.\n(A) Creating art, music, or stories.\n(S) Volunteering or spending time with friends and family.\n(E) Negotiating deals or coming up with business ideas.\n(C) Keeping things organized and tidy.",
        "My favorite school subject is:\n(R) Shop class or woodshop.\n(I) Science or math.\n(A) Art, music, or drama.\n(S) Psychology, sociology, or history.\n(E) Business or marketing.\n(C) English or foreign language (grammar & structure).",
        "When working on a project, I prefer to:\n(R) Work independently and solve problems with my hands.\n(I) Gather data and analyze information.\n(A) Use my imagination and creativity.\n(S) Collaborate with others and help them succeed.\n(E) Take charge and lead the way.\n(C) Follow a clear plan and make sure everything is accurate.",
        "I would be most satisfied working in a job that allows me to:\n(R) Use my skills to build or fix things.\n(I) Solve mysteries and uncover new knowledge.\n(A) Express myself creatively and come up with new ideas.\n(S) Help others and make a positive impact.\n(E) Persuade others and achieve success.\n(C) Work with details and complete tasks accurately.",
        "I find it rewarding to:\n(R) See a tangible result of my work.\n(I) Understand the world around me better.\n(A) Create something beautiful or meaningful.\n(S) Build relationships and connect with others.\n(E) Take risks and achieve ambitious goals.\n(C) Maintain order and follow established procedures.",
        "I am naturally good at:\n(R) Working with my hands and fixing things.\n(I) Analyzing problems and finding solutions.\n(A) Thinking creatively and coming up with new ideas.\n(S) Communicating with others and building relationships.\n(E) Persuading others and taking initiative.\n(C) Following instructions and paying attention to detail.",
        "In the future, I hope to:\n(R) Work with a skilled trade or become an engineer.\n(I) Become a scientist, researcher, or doctor.\n(A) Work in a creative field like music, art, or design.\n(S) Become a teacher, counselor, or social worker.\n(E) Start my own business or work in sales.\n(C) Work in an administrative role or become an accountant.",
        "When reading a book or watching a movie, I am most drawn to:\n(R) Action movies or stories about survival.\n(I) Documentaries or mysteries.\n(A) Fictional stories with strong emotional themes.\n(S) Biographies or stories about helping others.\n(E) Business biographies or stories about overcoming challenges.\n(C) Historical fiction or stories with clear structures.",
        "I would describe myself as:\n(R) Practical and hands-on.\n(I) Analytical and curious.\n(A) Creative and expressive.\n(S) Helpful and outgoing.\n(E) Persuasive and ambitious.\n(C) Organized and detail-oriented."
        
    ],
    " Personality Questions": [
        "Personality Domain - I would rather spend my free time:\n(R) Building or fixing something.\n(I) Researching a topic that interests me.\n(A) Creating art, music, or writing.\n(S) Volunteering or socializing with friends.\n(E) Brainstorming business ideas or convincing others of something.\n(C) Organizing my space or completing a detailed task.",
        "When working on a project, I am most motivated by:\n(R) Seeing a tangible result of my efforts.\n(I) Understanding the problem and finding a solution.\n(A) Expressing myself creatively and coming up with something new.\n(S) Helping others and making a positive impact.\n(E) Achieving success and reaching my goals.\n(C) Following a clear plan and ensuring accuracy.",
        "I find it most rewarding to:\n(R) Use my skills to solve practical problems.\n(I) Learn and understand new things.\n(A) Come up with unique ideas and express myself creatively.\n(S) Build relationships and connect with others.\n(E) Persuade others and achieve results.\n(C) Maintain order and complete tasks efficiently.",
        "When faced with a challenge, I tend to:\n(R) Take a hands-on approach and figure it out myself.\n(I) Analyze the situation and gather information.\n(A) Think creatively and come up with new solutions.\n(S) Seek help from others or offer support.\n(E) Take charge and find a way to win.\n(C) Develop a plan and follow it step-by-step.",
        "In a work environment, I prefer to:\n(R) Work independently and use my hands.\n(I) Work on problems that require research and analysis.\n(A) Have the freedom to express myself creatively.\n(S) Work collaboratively and help others succeed.\n(E) Take on leadership roles and influence others.\n(C) Work in a structured environment with clear expectations.",
        "I am naturally good at:\n(R) Working with tools and fixing things.\n(I) Analyzing information and solving puzzles.\n(A) Thinking outside the box and coming up with new ideas.\n(S) Communicating with others and building rapport.\n(E) Leading and motivating others.\n(C) Following instructions and completing tasks accurately.",
        "I am most drawn to books or movies that are:\n(R) Action-packed or involve practical skills.\n(I) Mysteries, documentaries, or science fiction.\n(A) Creative, imaginative, or evoke emotions.\n(S) Biographies, stories about relationships, or inspiring tales.\n(E) Business-oriented, thrillers, or stories about overcoming challenges.\n(C) Historical fiction, well-structured narratives, or anything with a clear plot.",
        "In the future, I hope to have a career that allows me to:\n(R) Work with my hands and create something tangible.\n(I) Conduct research, analyze data, or solve problems.\n(A) Express myself creatively and use my imagination.\n(S) Help others and make a positive impact on the world.\n(E) Lead others, take risks, and achieve ambitious goals.\n(C) Work in a structured environment and complete tasks efficiently.",
        "I am most energized by:\n(R) Working with a physical object and seeing progress.\n(I) Learning new things and solving complex problems.\n(A) Creating something new and expressing myself creatively.\n(S) Helping others and connecting with them on a personal level.\n(E) Taking on challenges and overcoming obstacles.\n(C) Working in a calm and organized environment.",
        "You describe yourself as someone who is:\n(R) Practical and results-oriented.\n(I) Analytical and curious.\n(A) Creative and imaginative.\n(S) Helpful and compassionate.\n(E) Persuasive and ambitious.\n(C) Organized and detail-oriented."
        
    ],
    " Values Questions": [
        "Values and Work Ethics Domain - When choosing a career, it's important for me to:\n(R) Use my skills to solve practical problems and build things.\n(I) Learn new things and be intellectually challenged.\n(A) Express myself creatively and make something beautiful.\n(S) Help others and make a positive impact on the world.\n(E) Take risks and achieve ambitious goals.\n(C) Work in a structured environment and follow clear procedures.",
        "In a work environment, I value:\n(R) Working independently and seeing tangible results.\n(I) Analyzing information and solving complex problems.\n(A) Freedom to be creative and express myself.\n(S) Collaboration and helping others succeed.\n(E) Competition and achieving success.\n(C) Efficiency, accuracy, and clear expectations.",
        "I believe it's important to:\n(R) Work hard and be practical.\n(I) Be curious and ask questions.\n(A) Be imaginative and think outside the box.\n(S) Be compassionate and helpful.\n(E) Be ambitious and take charge.\n(C) Be organized and detail-oriented.",
        "I am motivated by:\n(R) Seeing a finished product or practical application of my work.\n(I) Understanding complex problems and finding solutions.\n(A) Creating something new and expressing my unique ideas.\n(S) Helping others and making a positive contribution.\n(E) Achieving success and recognition.\n(C) Completing tasks accurately and efficiently.",
        "I enjoy working on tasks that are:\n(R) Hands-on and require practical skills.\n(I) Data-driven and require research and analysis.\n(A) Open-ended and allow for creative expression.\n(S) Collaborative and involve working with people.\n(E) Competitive and require strategic thinking.\n(C) Structured and have clear goals and procedures.",
        "I would rather work in a job that is:\n(R) Stable and predictable.\n(I) Challenging and intellectually stimulating.\n(A) Unstructured and allows for creativity.\n(S) Socially interactive and allows for helping others.\n(E) Fast-paced and competitive.\n(C) Organized and well-defined.",
        "When faced with a problem at work, I prefer to:\n(R) Find a practical solution and fix it myself.\n(I) Analyze the problem and gather data before acting.\n(A) Come up with a creative solution.\n(S) Discuss the problem with colleagues and find a collaborative solution.\n(E) Develop a strategy and take charge of finding a solution.\n(C) Follow existing procedures to solve the problem.",
        "I am most satisfied when my work:\n(R) Creates a tangible product or solves a practical problem.\n(I) Contributes to new knowledge or understanding.\n(A) Allows me to express myself creatively.\n(S) Helps others and makes a positive difference.\n(E) Leads to success and accomplishment.\n(C) Is accurate, efficient, and follows established procedures.",
        "I find it important to work for a company that values:\n(R) Quality craftsmanship and practicality.\n(I) Innovation and research.\n(A) Creativity and artistic expression.\n(S) Helping others and social responsibility.\n(E) Achievement and success.\n(C) Efficiency, organization, and clear communication.",
        "I am most stressed by work environments that are:\n(R) Disorganized and lack clear goals. (Opposite of Conventional)\n(I) Chaotic and lack intellectual stimulation. (Opposite of Investigative)\n(A) Restrictive and stifle creativity. (Opposite of Artistic)\n(S) Uncaring and lack opportunities to help others. (Opposite of Social)\n(E) Lacking in competition and challenge. (Opposite of Enterprising)\n(C) Unstructured and unpredictable. (Opposite of Conventional)",
        "I am most likely to leave a job because it doesn't provide opportunities to:\n(R) Use my skills and work with my hands. (Realistic)\n(I) Learn new things and solve problems. (Investigative)\n(A) Be creative and express myself. (Artistic)\n(S) Help others and make a positive impact. (Social)\n(E) Achieve success and take on challenges. (Enterprising)\n(C) Work in an organized and efficient manner. (Conventional)",
        "The most important quality in a leader is:\n(R) Practicality and problem-solving skills. (Realistic)\n(I) Curiosity and a thirst for knowledge. (Investigative)\n(A) Creativity and the ability to inspire others. (Artistic)\n(S) Empathy and a desire to help others succeed. (Social)\n(E) Vision, ambition, and the ability to drive results. (Enterprising)\n(C) Organization, attention to detail, and clear communication."
        ]
}

career_scopes = {
    'R': "Realistic: Carpenter, Chef/Baker, Agricultural Technician, Landscaper, Hairdresser, Electrical/Civil Engineer, Bus driver, and Construction worker",
    'I': "Investigative: Scientist, Researcher, Private Investigator, Data Analyst, Journalist, Software Developer, Economist, Psychiatrist, and Soil & Plant Scientist",
    'A': "Artistic: Creative Writer, Visual Artist, Digital Designer, Multi-media Animator, Video Producer, Actor, Photographer, and Landscape Architect",
    'S': "Social: Teacher, Counsellor, Public Relation Specialist, Nurse, Tour Guide, Waiter, Marketing Specialist, Fitness Trainer, and Social Worker",
    'E': "Enterprising: Business Owner, Real Estate Agent, Chef, Telemarketer, Advertising Specialist, Product Promoter, Insurance Sales Agent, and Lawyer",
    'C': "Conventional: Accountant, Auditor, Cashier, Receptionist, Dental Assistant, Clerk, Secretary, Web Developer, and Computer Security Specialist"
}


def calculate_percentages(responses):
    counts = {key: 0 for key in 'RIASCE'}
    total_responses = len(responses)

    for response in responses.values():
        if response in counts:
            counts[response] += 1

    percentages = {key: (value / total_responses) * 100 for key, value in counts.items()}
    return percentages



def get_summary(percentages, category):
    highest_percentage = max(percentages, key=percentages.get)
    representation = {
        'R': "Realistic - Practical, hands-on problems and solutions",
        'I': "Investigative - Working with ideas, thinking, and figuring things out",
        'A': "Artistic - Working with forms, designs, and patterns",
        'S': "Social - Working with others to help them learn and grow",
        'E': "Enterprising - Starting up and carrying out projects",
        'C': "Conventional - Following set procedures and routines"
    }

    summary = f"Your dominant area of interest in {category} is {highest_percentage}.\n"
    summary += f"This represents the '{highest_percentage}' type, which means: {representation[highest_percentage]}"
    summary += f"\n\nCareer options for {highest_percentage}: {career_scopes[highest_percentage]}"
    return summary



def write_to_excel(user_name, responses, summary):
    """Writes chatbot data to a new Excel sheet."""

    # Create a new workbook
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Set sheet title
    sheet.title = f"{user_name}'s Career Quiz Results"

    # Write headers
    sheet.cell(row=1, column=1).value = "Question"
    sheet.cell(row=1, column=2).value = "Response"

    # Write questions and responses
    row_num = 2
    for question, response in responses.items():
        sheet.cell(row=row_num, column=1).value = question
        sheet.cell(row=row_num, column=2).value = response
        row_num += 1

    # Write summary below the responses
    sheet.cell(row=row_num + 2, column=1).value = "Summary"
    sheet.cell(row=row_num + 2, column=2).value = summary

    # Save the workbook with a filename based on username
    wb.save(f"{user_name}_career_results.xlsx")



def chatbot():
    # Get the user's name
    user_name = input("Welcome to the Career Interest Quiz! What's your name? ")
    print(f"Hello, {user_name}! Let's start the quiz. Please answer each question with R, I, A, S, E, or C.")

    # Dictionary to store responses
    responses = {}

    # Ask each question categorically and get the user's response
    for category, questions_list in questions.items():
        print(f"\n{category}:")

        for i, question in enumerate(questions_list):
            print(f"\nQ: {i+1} in {category}:")
            print(question)
            response = input("Your answer (R/I/A/S/E/C): ").strip().upper()

            # Validate the response (same as before)

            # Store the response
            responses[f"Q: {i+1} in {category}"] = response

    # Calculate percentages for each category
    interest_responses = {k: v for k, v in responses.items() if 'Interst Questions' in k}
    interest_percentages = calculate_percentages(interest_responses)
    interest_summary = get_summary(interest_percentages, "Interest Questions")

    personality_responses = {k: v for k, v in responses.items() if 'Personality Questions' in k}
    personality_percentages = calculate_percentages(personality_responses)
    personality_summary = get_summary(personality_percentages, "Personality Questions")

    values_responses = {k: v for k, v in responses.items() if 'Values Questions' in k}
    values_percentages = calculate_percentages(values_responses)
    values_summary = get_summary(values_percentages, "Values Questions")

    # Print all summaries together
    print("\nSummary for Interest Questions:")

    print(interest_summary)

    print("\nSummary for Personality Questions:")

    print(personality_summary)

    print("\nSummary for Values Questions:")

    print(values_summary)

    # Write data to Excel sheet with overall summary
    overall_percentages = calculate_percentages(responses)
    overall_summary = get_summary(overall_percentages, "Overall Categories")
    write_to_excel(user_name, responses, overall_summary)

    # Print final message
    
    
    
    
    print(overall_summary)
    print("\nThank you for completing the quiz! Your results have been saved in an Excel file.")


# Run the chatbot
chatbot()
